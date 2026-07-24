"""Sinh data-catalog kiểm soát schema thật GSM (gsm-data-prod) → CSV + XLSX.

CSV = canonical (diffable, stdlib, luôn sinh). XLSX = deliverable Excel cho Cường
(cần openpyxl: `uv sync --extra catalog`). Nguồn sự thật = CATALOG dưới đây (Python
list-of-dicts, sạch, review git dễ). 13 bảng khớp ảnh metadata Cường gửi 2026-07-24.

Chạy:  uv run --extra catalog python scripts/build_data_catalog.py
       (không có openpyxl → chỉ sinh CSV, in cảnh báo)
"""

from __future__ import annotations

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "docs" / "data-catalog"

# Cột: 5 cột GỐC của Cường + 5 cột ta thêm (mapping/PII/availability/consumer/mockgen).
COLUMNS = [
    "full_path", "fields_known", "reason_usage", "priority", "usecase",
    "our_layer", "pii_columns", "availability", "consumed_by", "mockgen_strategy",
]

# 13 bảng — full_path, danh sách trường (nếu biết), + annotation của ta.
CATALOG: list[dict] = [
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_statistic_daily",
        "fields_known": "local_date, driver_id, completed_count, accepted_count, cancelled_count, "
        "total_request_calculate_complete, total_request_calculate_cancel, total_request_calculate_accept, "
        "count_cancel_not_relate_driver, total_rating, total_order_rating, count_rating_5_star, "
        "acceptance_rate, fulfillment_rate, cancellation_rate",
        "reason_usage": "Snapshot chỉ số vận doanh daily của tài xế",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC1 Phân tích hiệu suất; UC4 Đánh giá KPI",
        "our_layer": "L1' measured — daily KPI aggregate (thay/bổ sung DriverDayState)",
        "pii_columns": "driver_id (pseudonymize)",
        "availability": "đủ 15 cột",
        "consumed_by": "S1 bonus_feasibility, S3 f3_patterns, penalty-explain(UC6)",
        "mockgen_strategy": "aggregate từ sim events; ràng buộc rate=count/request (acceptance=accepted/req…)",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_online_hours_sap_id",
        "fields_known": "local_date, schedule_date, driver_id, full_name, sap_profile_id, hub_id, "
        "depot_id, phone_number, driver_type, online_time",
        "reason_usage": "Xác định số giờ online chính xác theo ngày/kỳ",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC1 Phân tích hiệu suất",
        "our_layer": "L1' — online hours (nuôi quỹ giờ/shift)",
        "pii_columns": "full_name, phone_number, sap_profile_id (DROP/hash)",
        "availability": "đủ 10 cột",
        "consumed_by": "S2 shift_dp, S5 weekly_khoan",
        "mockgen_strategy": "online_time từ go_online/offline sim; driver_type/depot từ profile",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_orders_rush_hours",
        "fields_known": "driver_id, local_date, total_order, commission, total_fee, revenue_not_relate_driver, "
        "total_order_normal_hour, commission_normal_hour, total_fee_normal_hour, revenue_not_relate_driver_normal_hour, "
        "total_order_rush_hour, commission_rush_hour, total_fee_rush_hour, revenue_not_relate_driver_rush_hour",
        "reason_usage": "Phân tích khung giờ chạy hiệu quả/kém",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC2 Phân tích điểm cần cải thiện",
        "our_layer": "L1' — revenue split rush/normal hour",
        "pii_columns": "driver_id",
        "availability": "đủ 14 cột",
        "consumed_by": "S2 shift_dp, S3",
        "mockgen_strategy": "split trips theo khung rush; commission=gross×share; ràng buộc normal+rush=total",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_bike_stoppoints",
        "fields_known": "driver_id, local_date, total_stoppoints, total_stoppoints_rush_hour",
        "reason_usage": "Xác định thói quen dừng đứng của tài xế",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC2 Phân tích điểm cần cải thiện",
        "our_layer": "L1' — stoppoints (idle proxy)",
        "pii_columns": "driver_id",
        "availability": "đủ 4 cột",
        "consumed_by": "S3, idle-reduction(UC5)",
        "mockgen_strategy": "đếm từ hex stay/idle segments sim; rush ⊆ total",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.kpi_driver_platform_calculator_gbq",
        "fields_known": "id, driver_id, driver_name, sap_id, status, week_key, week_start, week_end, "
        "kpi_month, kpi_year, email, tel, engname, depot_code, depot_name, vehicle_vin_number, "
        "vehicle_license_plate, vehicle_model, country, type, last_updated_date",
        "reason_usage": "Tính khoảng thiếu KPI target và thưởng",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC3 Theo dõi tiến độ KPI/thưởng",
        "our_layer": "L1' weekly KPI calculator — NỀN S5 weekly_khoan",
        "pii_columns": "driver_name, sap_id, email, tel, vehicle_vin_number, vehicle_license_plate (DROP/hash)",
        "availability": "đủ 21 cột; số TARGET/threshold cụ thể có thể ở meta/khác",
        "consumed_by": "S5 weekly_khoan",
        "mockgen_strategy": "weekly rollup; vehicle/depot từ profile; target/mốc = policy_bundle config (versioned)",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_income_daily",
        "fields_known": "driver_id, order_date, commission, total_order, total_fee, "
        "revenue_not_relate_driver, avg_daily_revenue, total_core_order",
        "reason_usage": "Theo dõi xu hướng thu nhập daily",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC3 Tiến độ KPI; UC4 Khả năng đạt KPI",
        "our_layer": "L1' — income daily (payout breakdown)",
        "pii_columns": "driver_id",
        "availability": "đủ 8 cột",
        "consumed_by": "S1, S5, F3 payout_breakdown",
        "mockgen_strategy": "commission=gross×share; total_core_order ⊆ total_order; avg=total_fee/total_order",
    },
    {
        "full_path": "gsm-data-prod.GSM_ORDER_DISPATCH_SERVICE_APPEND.trips",
        "fields_known": "CHƯA CÓ CỘT (cacbangdabiet.xlsx)",
        "reason_usage": "Tính toán mật độ cuốc thực tế theo vùng",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC5 Gợi ý hành vi chạy xe (Reduce Idle)",
        "our_layer": "L1' trip-level (~ TripRecord) — nền demand density",
        "pii_columns": "driver_id, customer_id, lat/lon (hex-agg)",
        "availability": "THIẾU CỘT → infer từ TripRecord ta + XIN GSM",
        "consumed_by": "S4 capacity_alloc, demand density, idle(UC5)",
        "mockgen_strategy": "tái dùng shape trip_record; aggregate ra mật độ theo hex×bucket",
    },
    {
        "full_path": "gsm-data-prod.GSM_MISSION_SERVICE_APPEND.public_driver_hex_tracking",
        "fields_known": "id, driver_id, campaign_id, log_id, init_hex, current_hex, last_hex, target_hex, "
        "last_seen_at, entered_current_hex_at, stay_duration_seconds, reached_target, reached_target_at, "
        "hex_history, created_at, updated_at, schedule_job_id, datastream_metadata, tracking_status",
        "reason_usage": "Xác định vị trí hiện tại và lịch sử hex zone",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC5 Reduce Idle",
        "our_layer": "L1' hex movement (~ GPSPing aggregated to H3)",
        "pii_columns": "driver_id; hex (coarse OK)",
        "availability": "đủ 19 cột",
        "consumed_by": "idle-reduction(UC5), S4",
        "mockgen_strategy": "từ chuỗi GPS→H3 sim; stay_duration từ idle; target_hex từ reposition mission",
    },
    {
        "full_path": "gsm-data-prod.M_DRIVER_KPI_REWARD.driver_penalization_ATA",
        "fields_known": "CHƯA CÓ CỘT (cacbangdabiet.xlsx)",
        "reason_usage": "Giải thích nguyên nhân hiệu suất thấp & rủi ro bị phạt",
        "priority": "Ưu tiên cao nhất (KPI)",
        "usecase": "UC6 Giải thích hiệu suất; UC8 Tránh vi phạm",
        "our_layer": "L1' penalization events (clawback/deduction)",
        "pii_columns": "driver_id",
        "availability": "THIẾU CỘT → infer {driver_id, date, penalty_type, amount_vnd, reason, week} + XIN",
        "consumed_by": "penalty-explain(UC6), F3",
        "mockgen_strategy": "rule-based từ vi phạm conduct + clawback khoán tuần (hiếm)",
    },
    {
        "full_path": "gsm-data-prod.M_BROADCASTING_SERVICE_APPEND.public_frauds",
        "fields_known": "CHƯA CÓ CỘT (cacbangdabiet.xlsx)",
        "reason_usage": "Bắn cảnh báo rủi ro vi phạm hành vi",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC7 Cảnh báo lệch route/bất thường",
        "our_layer": "L1' fraud/anomaly flags",
        "pii_columns": "driver_id",
        "availability": "THIẾU CỘT → infer {driver_id, t, fraud_type, severity, evidence_ref, status} + XIN",
        "consumed_by": "anomaly-alert(UC7)",
        "mockgen_strategy": "flag hiếm rule-based (lệch route/bất thường) — nhãn INFERRED, không kết tội",
    },
    {
        "full_path": "gsm-data-prod.GSM_MISSION_SERVICE_APPEND.public_user_mission_progress",
        "fields_known": "CHƯA CÓ CỘT (cacbangdabiet.xlsx)",
        "reason_usage": "Gợi ý mini task tăng thu nhập bổ sung",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC8 Hướng dẫn tránh bị trừ tiền / mini task",
        "our_layer": "L1' mission progress (per driver × mission)",
        "pii_columns": "driver_id",
        "availability": "THIẾU CỘT → infer {id, driver_id, mission_id, progress, target, state, updated_at} + XIN",
        "consumed_by": "S6 mission_knapsack(UC8)",
        "mockgen_strategy": "progress đối chiếu target trong public_mission; nhất quán với earn_history",
    },
    {
        "full_path": "gsm-data-prod.GSM_MISSION_SERVICE_APPEND.public_mission",
        "fields_known": "id, created_at, updated_at, deleted_at, created_by, updated_by, mission_type, "
        "parent_id, name, state, audience, description, start_time, end_time, point_id, rewards, "
        "mission_claim, mission_code, time_claim_reward, rule_code, meta_data, contract_type, "
        "qualify_execute_code, status, datastream_metadata, business_code, show_only, is_ddi_mission",
        "reason_usage": "Nhiệm vụ/mini task theo khung giờ",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC8 mini task theo khung giờ",
        "our_layer": "L0/L1' mission catalog (reference, time-windowed)",
        "pii_columns": "(catalog — không PII driver)",
        "availability": "đủ 28 cột",
        "consumed_by": "S6 mission_knapsack",
        "mockgen_strategy": "catalog mission: rewards, start/end_time (khung giờ), rule_code, audience",
    },
    {
        "full_path": "gsm-data-prod.GSM_MISSION_SERVICE_APPEND.public_mission_earn_history",
        "fields_known": "id, created_at, updated_at, deleted_at, mission_id, order_id, order_status, "
        "driver_id, customer_id, service_type, order_time, complete_time, travel_mode, sap_contract_type, "
        "type, count_order, count_stoppoint, earn, description, datastream_metadata, reward_level",
        "reason_usage": "Lịch sử nhận thưởng nhiệm vụ của tài xế",
        "priority": "Ưu tiên cao nhất (APPEND)",
        "usecase": "UC3 Tiến độ KPI/thưởng; UC8 Mini task",
        "our_layer": "L1' mission earn history (event)",
        "pii_columns": "driver_id, customer_id (DROP customer_id)",
        "availability": "đủ 21 cột",
        "consumed_by": "S6 mission_knapsack, income breakdown",
        "mockgen_strategy": "từ mission hoàn thành × trips; earn từ public_mission.rewards; nhất quán progress",
    },
]


def write_csv(path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for row in CATALOG:
            w.writerow(row)


def write_xlsx(path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "gsm-data-catalog"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for row in CATALOG:
        ws.append([row[c] for c in COLUMNS])
    widths = {"full_path": 55, "fields_known": 60, "reason_usage": 30, "priority": 22,
              "usecase": 28, "our_layer": 40, "pii_columns": 32, "availability": 34,
              "consumed_by": 32, "mockgen_strategy": 55}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[chr(64 + i)].width = widths.get(col, 24)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)
    return True


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    assert len(CATALOG) == 13, f"phải đúng 13 bảng, đang có {len(CATALOG)}"
    csv_path = OUT_DIR / "gsm-data-catalog.csv"
    write_csv(csv_path)
    print(f"CSV written ({len(CATALOG)} tables): {csv_path}")
    xlsx_path = OUT_DIR / "gsm-data-catalog.xlsx"
    if write_xlsx(xlsx_path):
        print(f"XLSX written: {xlsx_path}")
    else:
        print("XLSX skipped (no openpyxl) - run `uv sync --extra catalog` then retry.")


if __name__ == "__main__":
    main()
